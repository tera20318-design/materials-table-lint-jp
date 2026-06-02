from __future__ import annotations

import csv
from pathlib import Path

from .models import Table


def read_table(path: Path, sheet: str | None = None) -> Table:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path, sheet=sheet)
    raise ValueError(f"unsupported input extension: {suffix}")


def read_csv(path: Path) -> Table:
    metadata: dict[str, str] = {}
    data_lines: list[str] = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            key, sep, value = stripped[1:].partition("=")
            if sep:
                metadata[key.strip()] = value.strip()
            continue
        if stripped:
            data_lines.append(line)
    if not data_lines:
        raise ValueError("CSV has no table header")
    reader = csv.DictReader(data_lines)
    if not reader.fieldnames:
        raise ValueError("CSV has no table header")
    headers = tuple(str(item) for item in reader.fieldnames)
    rows = tuple(
        {key: value or "" for key, value in row.items() if key is not None} for row in reader
    )
    return Table(path=path, metadata=metadata, headers=headers, rows=rows)


def read_xlsx(path: Path, sheet: str | None = None) -> Table:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel input requires optional dependency: "
            "python -m pip install materials-table-lint-jp[xlsx]"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    metadata: dict[str, str] = {}
    header: tuple[str, ...] | None = None
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(values_only=True):
        cells = ["" if value is None else str(value).strip() for value in values]
        if not any(cells):
            continue
        first = cells[0]
        if first.startswith("#"):
            key, sep, value = first[1:].partition("=")
            if sep:
                metadata[key.strip()] = value.strip()
            continue
        if header is None:
            header = tuple(cells)
            continue
        row = {
            column: cells[index] if index < len(cells) else ""
            for index, column in enumerate(header)
        }
        rows.append(row)
    if header is None:
        raise ValueError("XLSX has no table header")
    return Table(path=path, metadata=metadata, headers=header, rows=tuple(rows))
